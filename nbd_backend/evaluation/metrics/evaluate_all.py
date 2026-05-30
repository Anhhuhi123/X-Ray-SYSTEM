"""
evaluate_all.py
───────────────
Orchestrator: chạy tất cả 4 metrics và xuất báo cáo tổng hợp.

Metrics được chạy:
    1. Context Precision  — cosine similarity(question, contexts)
    2. Faithfulness       — LLM judge: answer supported by contexts?
    3. Answer Correctness — BERTScore F1 (answer vs ground_truth)
    4. Context Recall     — LLM judge: contexts cover ground_truth?

Output (cùng thư mục --output_dir):
    evaluation_report.json   — full per-item data
    evaluation_report.csv    — bảng phẳng cho spreadsheet / pandas
    evaluation_report.xlsx   — Excel với conditional formatting + sheet Summary

Bạn vẫn có thể chạy từng metric rời rạc:
    python metrics/evaluate_context_precision.py
    python metrics/evaluate_faithfulness.py
    python metrics/evaluate_answer_relevancy.py
    python metrics/evaluate_context_recall.py

Params:
    --input_file      : file JSON đã generate (default: results/140_evaluated_gemini.json)
    --output_dir      : thư mục chứa output   (default: metrics_output/)
    --language        : 'vi' hoặc 'en'
    --llm_model       : model cho RAGAS judge  (faithfulness + context_recall)
    --embed_model     : override sentence-transformers model (context_precision)
    --bertscore_model : override BERTScore model (answer_correctness)
    --skip_metrics    : bỏ qua metric(s) — vd: --skip_metrics faithfulness context_recall
    --no_ragas        : bỏ qua RAGAS/LLM, dùng heuristic token-overlap cho faithfulness
                        và context_recall (không cần API key, chạy local hoàn toàn)
"""

import argparse
import json
import logging
import os
import sys
from pathlib import Path
from typing import Any

from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ── Đảm bảo metrics package có thể import ────────────────────────────────────
_METRICS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(_METRICS_DIR))

# Default models
DEFAULT_GEMINI_MODEL = "gemini-2.5-flash"
DEFAULT_OPENAI_MODEL = "gpt-4o-mini"

DEFAULT_EMBED_MODELS = {
    "vi": "sentence-transformers/paraphrase-multilingual-mpnet-base-v2",
    "en": "sentence-transformers/all-mpnet-base-v2",
}
DEFAULT_BERTSCORE_MODELS = {
    "vi": "bert-base-multilingual-cased",
    "en": "roberta-large",
}

ALL_METRIC_NAMES = [
    "context_precision",
    "faithfulness",
    "answer_correctness",
    "context_recall",
]

# ── Thresholds cho conditional formatting trong Excel ────────────────────────
SCORE_THRESHOLDS = {
    "high": 0.75,  # ≥ 0.75 → xanh lá
    "mid": 0.50,  # ≥ 0.50 → vàng cam
    # < 0.50       → đỏ
}


# ──────────────────────────────────────────────────────────────────────────────
# Data helpers
# ──────────────────────────────────────────────────────────────────────────────
def load_data(input_file: str) -> list[dict[str, Any]]:
    path = Path(input_file)
    if not path.exists():
        print(f"[ERROR] Input file not found: {input_file}", file=sys.stderr)
        sys.exit(1)
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(output_file: str, data: Any) -> None:
    out_path = Path(output_file)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ──────────────────────────────────────────────────────────────────────────────
# Metric runner helpers
# ──────────────────────────────────────────────────────────────────────────────
def _run_context_precision(
    input_file: str,
    output_file: str,
    embed_model: str,
    language: str,
) -> dict[str, Any]:
    """Gọi compute_context_precision và trả về output dict."""
    from evaluate_context_precision import compute_context_precision  # type: ignore

    print("\n" + "═" * 60)
    print("▶  Running: Context Precision")
    print("═" * 60)
    compute_context_precision(
        input_file=input_file,
        output_file=output_file,
        model_name=embed_model,
        language=language,
    )
    with open(output_file, encoding="utf-8") as f:
        return json.load(f)


def _run_faithfulness(
    input_file: str,
    output_file: str,
    llm_model: str,
    language: str,
    use_ragas: bool,
) -> dict[str, Any]:
    """Gọi compute_faithfulness và trả về output dict."""
    from evaluate_faithfulness import compute_faithfulness  # type: ignore

    print("\n" + "═" * 60)
    print("▶  Running: Faithfulness" + ("" if use_ragas else "  [heuristic mode]"))
    print("═" * 60)
    compute_faithfulness(
        input_file=input_file,
        output_file=output_file,
        model_name=llm_model,
        language=language,
        use_ragas=use_ragas,
    )
    with open(output_file, encoding="utf-8") as f:
        return json.load(f)


def _run_answer_correctness(
    input_file: str,
    output_file: str,
    bertscore_model: str,
    language: str,
) -> dict[str, Any]:
    """Gọi compute_answer_relevancy (thực chất đo Answer Correctness) và trả về output dict."""
    from evaluate_answer_relevancy import compute_answer_relevancy  # type: ignore

    print("\n" + "═" * 60)
    print("▶  Running: Answer Correctness (BERTScore)")
    print("═" * 60)
    compute_answer_relevancy(
        input_file=input_file,
        output_file=output_file,
        model_name=bertscore_model,
        language=language,
    )
    with open(output_file, encoding="utf-8") as f:
        return json.load(f)


def _run_context_recall(
    input_file: str,
    output_file: str,
    llm_model: str,
    language: str,
    use_ragas: bool,
) -> dict[str, Any]:
    """Gọi compute_context_recall và trả về output dict."""
    from evaluate_context_recall import compute_context_recall  # type: ignore

    print("\n" + "═" * 60)
    print("▶  Running: Context Recall" + ("" if use_ragas else "  [heuristic mode]"))
    print("═" * 60)
    return compute_context_recall(
        input_file=input_file,
        output_file=output_file,
        model_name=llm_model,
        language=language,
        use_ragas=use_ragas,
    )


# ──────────────────────────────────────────────────────────────────────────────
# Merge per-item results
# ──────────────────────────────────────────────────────────────────────────────
def _merge_results(
    data: list[dict[str, Any]],
    cp_output: dict[str, Any] | None,
    faith_output: dict[str, Any] | None,
    ac_output: dict[str, Any] | None,
    cr_output: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    """
    Ghép per-item scores từ 4 metrics theo thứ tự question.
    Dùng index để match vì thứ tự phải bảo toàn.
    """
    cp_results = cp_output["results"] if cp_output else [{}] * len(data)
    faith_results = faith_output["results"] if faith_output else [{}] * len(data)
    ac_results = ac_output["results"] if ac_output else [{}] * len(data)
    cr_results = cr_output["results"] if cr_output else [{}] * len(data)

    merged = []
    for idx, item in enumerate(data):
        cp = cp_results[idx] if idx < len(cp_results) else {}
        fa = faith_results[idx] if idx < len(faith_results) else {}
        ac = ac_results[idx] if idx < len(ac_results) else {}
        cr = cr_results[idx] if idx < len(cr_results) else {}

        merged.append(
            {
                "no": idx + 1,
                "question": item.get("question", ""),
                # ── Context Precision ────────────────────────────────────────
                "context_precision": cp.get("context_precision_score"),
                # ── Faithfulness ─────────────────────────────────────────────
                "faithfulness": fa.get("faithfulness_score"),
                "faithfulness_method": fa.get("method"),
                # ── Answer Correctness (BERTScore F1) ────────────────────────
                "answer_correctness_f1": ac.get("bertscore_f1"),
                "answer_correctness_precision": ac.get("bertscore_precision"),
                "answer_correctness_recall": ac.get("bertscore_recall"),
                "answer_correctness_skipped": ac.get("skipped", False),
                # ── Context Recall ───────────────────────────────────────────
                "context_recall": cr.get("context_recall_score"),
                "context_recall_method": cr.get("method"),
            }
        )
    return merged


# ──────────────────────────────────────────────────────────────────────────────
# Summary stats
# ──────────────────────────────────────────────────────────────────────────────
def _compute_summary(
    merged: list[dict[str, Any]],
    skipped_metrics: list[str],
) -> dict[str, Any]:
    import numpy as np

    def _safe_mean(key: str) -> float | None:
        vals = [r[key] for r in merged if r.get(key) is not None]
        return round(float(np.mean(vals)), 4) if vals else None

    def _safe_std(key: str) -> float | None:
        vals = [r[key] for r in merged if r.get(key) is not None]
        return round(float(np.std(vals)), 4) if vals else None

    def _safe_min(key: str) -> float | None:
        vals = [r[key] for r in merged if r.get(key) is not None]
        return round(float(np.min(vals)), 4) if vals else None

    def _safe_max(key: str) -> float | None:
        vals = [r[key] for r in merged if r.get(key) is not None]
        return round(float(np.max(vals)), 4) if vals else None

    return {
        "num_items": len(merged),
        "skipped_metrics": skipped_metrics,
        "context_precision": {
            "mean": _safe_mean("context_precision"),
            "std": _safe_std("context_precision"),
            "min": _safe_min("context_precision"),
            "max": _safe_max("context_precision"),
        },
        "faithfulness": {
            "mean": _safe_mean("faithfulness"),
            "std": _safe_std("faithfulness"),
            "min": _safe_min("faithfulness"),
            "max": _safe_max("faithfulness"),
        },
        "answer_correctness_f1": {
            "mean": _safe_mean("answer_correctness_f1"),
            "std": _safe_std("answer_correctness_f1"),
            "min": _safe_min("answer_correctness_f1"),
            "max": _safe_max("answer_correctness_f1"),
        },
        "context_recall": {
            "mean": _safe_mean("context_recall"),
            "std": _safe_std("context_recall"),
            "min": _safe_min("context_recall"),
            "max": _safe_max("context_recall"),
        },
    }


# ──────────────────────────────────────────────────────────────────────────────
# Export CSV
# ──────────────────────────────────────────────────────────────────────────────
def _export_csv(merged: list[dict[str, Any]], csv_path: str) -> None:
    try:
        import pandas as pd
    except ImportError:
        logger.warning("pandas not installed — skipping CSV export. pip install pandas")
        return

    df = pd.DataFrame(merged)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"  📄  CSV  → {csv_path}")


# ──────────────────────────────────────────────────────────────────────────────
# Export Excel với conditional formatting
# ──────────────────────────────────────────────────────────────────────────────
def _export_excel(
    merged: list[dict[str, Any]],
    summary: dict[str, Any],
    xlsx_path: str,
) -> None:
    try:
        import pandas as pd
    except ImportError:
        logger.warning(
            "pandas not installed — skipping Excel export. pip install pandas"
        )
        return
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning(
            "openpyxl not installed — skipping Excel export. pip install openpyxl"
        )
        return

    # ── Màu sắc ──────────────────────────────────────────────────────────────
    COLOR_HIGH = "C6EFCE"  # xanh lá nhạt
    COLOR_MID = "FFEB9C"  # vàng nhạt
    COLOR_LOW = "FFC7CE"  # đỏ nhạt
    COLOR_HEADER = "2E4057"  # xanh đậm header
    COLOR_SUBHDR = "4A7C94"  # xanh nhạt sub-header
    COLOR_ALT = "F0F4F8"  # xám nhạt alternate row

    FILL_HIGH = PatternFill("solid", fgColor=COLOR_HIGH)
    FILL_MID = PatternFill("solid", fgColor=COLOR_MID)
    FILL_LOW = PatternFill("solid", fgColor=COLOR_LOW)
    FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
    FILL_SUBHDR = PatternFill("solid", fgColor=COLOR_SUBHDR)
    FILL_ALT = PatternFill("solid", fgColor=COLOR_ALT)

    FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
    FONT_BOLD = Font(bold=True, size=10)
    FONT_NORMAL = Font(size=10)

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _score_fill(val) -> PatternFill | None:
        if val is None:
            return None
        if val >= SCORE_THRESHOLDS["high"]:
            return FILL_HIGH
        if val >= SCORE_THRESHOLDS["mid"]:
            return FILL_MID
        return FILL_LOW

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1: Per-item Results
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Results"

    # Định nghĩa cột và nhãn
    score_cols = [
        "context_precision",
        "faithfulness",
        "answer_correctness_f1",
        "context_recall",
    ]
    col_labels = {
        "no": "#",
        "question": "Question",
        "context_precision": "Context Precision",
        "faithfulness": "Faithfulness",
        "answer_correctness_f1": "Answer Correctness (F1)",
        "context_recall": "Context Recall",
        "faithfulness_method": "Faithfulness Method",
        "context_recall_method": "Recall Method",
        "answer_correctness_skipped": "AC Skipped?",
    }
    columns = list(col_labels.keys())

    # Header row
    for col_idx, col_key in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_labels[col_key])
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, item in enumerate(merged, start=2):
        alt = row_idx % 2 == 0
        for col_idx, col_key in enumerate(columns, start=1):
            val = item.get(col_key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_key != "question" else "left",
                vertical="center",
                wrap_text=(col_key == "question"),
            )

            # Score cells: conditional formatting
            if col_key in score_cols:
                fill = _score_fill(val)
                if fill:
                    cell.fill = fill
            elif alt:
                cell.fill = FILL_ALT

        ws.row_dimensions[row_idx].height = 45

    # Column widths
    col_widths = {
        "no": 5,
        "question": 55,
        "context_precision": 18,
        "faithfulness": 15,
        "answer_correctness_f1": 22,
        "context_recall": 16,
        "faithfulness_method": 20,
        "context_recall_method": 20,
        "answer_correctness_skipped": 14,
    }
    for col_idx, col_key in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(
            col_key, 14
        )

    # Freeze header row
    ws.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2: Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")

    summary_rows = [
        ("Metric", "Mean", "Std", "Min", "Max"),
        (
            "Context Precision",
            summary["context_precision"]["mean"],
            summary["context_precision"]["std"],
            summary["context_precision"]["min"],
            summary["context_precision"]["max"],
        ),
        (
            "Faithfulness",
            summary["faithfulness"]["mean"],
            summary["faithfulness"]["std"],
            summary["faithfulness"]["min"],
            summary["faithfulness"]["max"],
        ),
        (
            "Answer Correctness (F1)",
            summary["answer_correctness_f1"]["mean"],
            summary["answer_correctness_f1"]["std"],
            summary["answer_correctness_f1"]["min"],
            summary["answer_correctness_f1"]["max"],
        ),
        (
            "Context Recall",
            summary["context_recall"]["mean"],
            summary["context_recall"]["std"],
            summary["context_recall"]["min"],
            summary["context_recall"]["max"],
        ),
    ]

    for r_idx, row in enumerate(summary_rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if r_idx == 1:  # Header
                cell.fill = FILL_SUBHDR
                cell.font = FONT_HEADER
            else:
                cell.font = FONT_BOLD if c_idx == 1 else FONT_NORMAL
                # Conditional coloring trên cột Mean
                if c_idx == 2 and isinstance(val, float):
                    fill = _score_fill(val)
                    if fill:
                        cell.fill = fill

    # Extra info
    ws2.cell(row=len(summary_rows) + 2, column=1, value="Total Items").font = FONT_BOLD
    ws2.cell(row=len(summary_rows) + 2, column=2, value=summary["num_items"])
    if summary["skipped_metrics"]:
        ws2.cell(
            row=len(summary_rows) + 3, column=1, value="Skipped Metrics"
        ).font = FONT_BOLD
        ws2.cell(
            row=len(summary_rows) + 3,
            column=2,
            value=", ".join(summary["skipped_metrics"]),
        )

    # Column widths
    ws2.column_dimensions["A"].width = 28
    for letter in ["B", "C", "D", "E"]:
        ws2.column_dimensions[letter].width = 14
    ws2.row_dimensions[1].height = 25

    # ── Legend ───────────────────────────────────────────────────────────────
    legend_start = len(summary_rows) + 5
    ws2.cell(row=legend_start, column=1, value="Legend").font = FONT_BOLD
    legends = [
        (FILL_HIGH, f"≥ {SCORE_THRESHOLDS['high']}  — Good"),
        (FILL_MID, f"≥ {SCORE_THRESHOLDS['mid']}  — Medium"),
        (FILL_LOW, f"<  {SCORE_THRESHOLDS['mid']}  — Low"),
    ]
    for i, (fill, label) in enumerate(legends):
        r = legend_start + 1 + i
        color_cell = ws2.cell(row=r, column=1)
        color_cell.fill = fill
        color_cell.border = THIN_BORDER
        color_cell.alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=2, value=label).font = FONT_NORMAL

    # ── Save ─────────────────────────────────────────────────────────────────
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"  📊  XLSX → {xlsx_path}")


# ──────────────────────────────────────────────────────────────────────────────
# Print summary table to stdout
# ──────────────────────────────────────────────────────────────────────────────
def _print_summary_table(summary: dict[str, Any]) -> None:
    metrics = [
        ("Context Precision", "context_precision"),
        ("Faithfulness", "faithfulness"),
        ("Answer Correctness (F1)", "answer_correctness_f1"),
        ("Context Recall", "context_recall"),
    ]

    print("\n" + "═" * 72)
    print("  EVALUATION REPORT — SUMMARY")
    print("═" * 72)
    print(f"  {'Metric':<28} {'Mean':>8}  {'Std':>7}  {'Min':>7}  {'Max':>7}")
    print("  " + "─" * 68)

    for label, key in metrics:
        s = summary[key]
        mean = f"{s['mean']:.4f}" if s["mean"] is not None else "  N/A "
        std = f"{s['std']:.4f}" if s["std"] is not None else "  N/A "
        mn = f"{s['min']:.4f}" if s["min"] is not None else "  N/A "
        mx = f"{s['max']:.4f}" if s["max"] is not None else "  N/A "
        print(f"  {label:<28} {mean:>8}  {std:>7}  {mn:>7}  {mx:>7}")

    print("═" * 72)
    print(f"  Total items: {summary['num_items']}")
    if summary["skipped_metrics"]:
        print(f"  Skipped:     {', '.join(summary['skipped_metrics'])}")
    print()


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────
def evaluate_all(
    input_file: str,
    output_dir: str,
    language: str,
    llm_model: str,
    embed_model: str,
    bertscore_model: str,
    skip_metrics: list[str],
    use_ragas: bool = True,
) -> None:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    data = load_data(input_file)
    print(f"\nLoaded {len(data)} items from: {input_file}")
    print(f"Metrics to run: {[m for m in ALL_METRIC_NAMES if m not in skip_metrics]}")
    if skip_metrics:
        print(f"Skipping      : {skip_metrics}")
    print(f"RAGAS/LLM     : {'enabled' if use_ragas else 'disabled (heuristic mode)'}")

    # ── Run metrics ───────────────────────────────────────────────────────────
    cp_output = None
    faith_output = None
    ac_output = None
    cr_output = None

    if "context_precision" not in skip_metrics:
        cp_output = _run_context_precision(
            input_file=input_file,
            output_file=str(out / "context_precision_scores.json"),
            embed_model=embed_model,
            language=language,
        )

    if "faithfulness" not in skip_metrics:
        faith_output = _run_faithfulness(
            input_file=input_file,
            output_file=str(out / "faithfulness_scores.json"),
            llm_model=llm_model,
            language=language,
            use_ragas=use_ragas,
        )

    if "answer_correctness" not in skip_metrics:
        ac_output = _run_answer_correctness(
            input_file=input_file,
            output_file=str(out / "answer_correctness_scores.json"),
            bertscore_model=bertscore_model,
            language=language,
        )

    if "context_recall" not in skip_metrics:
        cr_output = _run_context_recall(
            input_file=input_file,
            output_file=str(out / "context_recall_scores.json"),
            llm_model=llm_model,
            language=language,
            use_ragas=use_ragas,
        )

    # ── Merge & summarize ────────────────────────────────────────────────────
    merged = _merge_results(data, cp_output, faith_output, ac_output, cr_output)
    summary = _compute_summary(merged, skip_metrics)

    full_report = {
        "summary": summary,
        "config": {
            "input_file": input_file,
            "language": language,
            "llm_model": llm_model,
            "embed_model": embed_model,
            "bertscore_model": bertscore_model,
            "skipped_metrics": skip_metrics,
            "use_ragas": use_ragas,
        },
        "results": merged,
    }

    # ── Save outputs ─────────────────────────────────────────────────────────
    json_path = str(out / "evaluation_report.json")
    csv_path = str(out / "evaluation_report.csv")
    xlsx_path = str(out / "evaluation_report.xlsx")

    save_json(json_path, full_report)
    print(f"\n  📋  JSON → {json_path}")

    _export_csv(merged, csv_path)
    _export_excel(merged, summary, xlsx_path)

    # ── Print summary ─────────────────────────────────────────────────────────
    _print_summary_table(summary)
    print("✅  All done!")


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────
def main() -> None:
    backend_dir = Path(__file__).resolve().parent.parent.parent
    default_input = str(
        backend_dir / "evaluation" / "results" / "140_evaluated_gemini.json"
    )
    default_output_dir = str(backend_dir / "evaluation" / "metrics_output")

    # Auto-detect default LLM
    if os.environ.get("OPENAI_API_KEY"):
        default_llm = DEFAULT_OPENAI_MODEL
    else:
        default_llm = DEFAULT_GEMINI_MODEL

    parser = argparse.ArgumentParser(
        description=(
            "Run all 4 RAG evaluation metrics and produce a combined report "
            "(JSON + CSV + Excel). Each metric can still be run independently."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Run all metrics (default settings, language=vi)
  python metrics/evaluate_all.py

  # Run with OpenAI judge, custom input
  python metrics/evaluate_all.py --input_file results/140_evaluated_openai.json --llm_model gpt-4o

  # Skip LLM-heavy metrics during quick testing
  python metrics/evaluate_all.py --skip_metrics faithfulness context_recall

  # Full end-to-end test without any API/RAGAS dependency (heuristic only)
  python metrics/evaluate_all.py --no_ragas
        """,
    )
    parser.add_argument(
        "--input_file",
        type=str,
        default=default_input,
        help="Path to evaluated results JSON. Default: %(default)s",
    )
    parser.add_argument(
        "--output_dir",
        type=str,
        default=default_output_dir,
        help="Directory to save all output files. Default: %(default)s",
    )
    parser.add_argument(
        "--language",
        type=str,
        choices=["vi", "en"],
        default="vi",
        help="Language of the data. Affects default embedding/BERTScore model. Default: %(default)s",
    )
    parser.add_argument(
        "--llm_model",
        type=str,
        default=default_llm,
        help=(
            "LLM model for RAGAS judge (Faithfulness + Context Recall). "
            f"Default: {default_llm} (based on available API keys)"
        ),
    )
    parser.add_argument(
        "--embed_model",
        type=str,
        default=None,
        help=(
            "Override sentence-transformers model for Context Precision. "
            "Default: auto-selected by --language"
        ),
    )
    parser.add_argument(
        "--bertscore_model",
        type=str,
        default=None,
        help=(
            "Override BERTScore model for Answer Correctness. "
            "Default: auto-selected by --language"
        ),
    )
    parser.add_argument(
        "--skip_metrics",
        nargs="*",
        choices=ALL_METRIC_NAMES,
        default=[],
        metavar="METRIC",
        help=(
            "Space-separated list of metrics to skip. "
            f"Choices: {ALL_METRIC_NAMES}. "
            "Example: --skip_metrics faithfulness context_recall"
        ),
    )
    use_ragas_group = parser.add_mutually_exclusive_group()
    use_ragas_group.add_argument(
        "--use_ragas",
        dest="use_ragas",
        action="store_true",
        help="Enable RAGAS/LLM judge for Faithfulness and Context Recall. Default: enabled.",
    )
    use_ragas_group.add_argument(
        "--no_ragas",
        dest="use_ragas",
        action="store_false",
        help=(
            "Disable RAGAS/LLM for Faithfulness and Context Recall. "
            "Uses fast token-overlap heuristic instead — no API key needed. "
            "Useful for end-to-end smoke tests."
        ),
    )
    parser.set_defaults(use_ragas=True)

    args = parser.parse_args()

    # Resolve model names
    language = args.language
    embed_model = args.embed_model or DEFAULT_EMBED_MODELS[language]
    bert_model = args.bertscore_model or DEFAULT_BERTSCORE_MODELS[language]
    skip_metrics = args.skip_metrics or []

    evaluate_all(
        input_file=args.input_file,
        output_dir=args.output_dir,
        language=language,
        llm_model=args.llm_model,
        embed_model=embed_model,
        bertscore_model=bert_model,
        skip_metrics=skip_metrics,
        use_ragas=args.use_ragas,
    )


if __name__ == "__main__":
    main()
